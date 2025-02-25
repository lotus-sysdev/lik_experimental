# Standard Library Imports
import os
import re
import csv
import json

# Third-Party Imports
import requests
import openpyxl
from PIL import Image
from openpyxl_image_loader import SheetImageLoader
from datetime import datetime, timedelta
# Local Imports
from .models import *
from .forms import *
from .decorators import *

# Django Imports
from django.urls import reverse
from django.conf import settings
from django.utils import timezone
from django.contrib import messages
from django.core import serializers
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.db.models import Prefetch, Count, Q
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.core.files.uploadedfile import SimpleUploadedFile
from django.shortcuts import redirect, render, get_object_or_404
from django.db.models.functions import ExtractDay, ExtractMonth, ExtractYear, Upper
from datetime import datetime, timedelta


# -------------------- Placeholder for homepage --------------------#
@login_required
def home(request):
    form = DashboardFilterForm(request.GET)
    
    if form.is_valid():
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')

        items = Items.objects.all()
        if start_date and end_date:
            items = items.filter(tanggal_pemesanan__range=[start_date, end_date])

        item_counts = items.count()
        approved_item_counts = items.filter(is_approved=True).count()
        grouped_counts = items.values('customer__nama_pt', 'pic__nama').annotate(count=Count('SKU')).order_by('-count', 'customer__nama_pt', 'pic__nama')
        category_counts = items.values('category__name').annotate(count=Count('SKU')).order_by('-count', 'category__name')
        sku_counts = items.values(
            day=ExtractDay('tanggal_pemesanan'),
            month=ExtractMonth('tanggal_pemesanan'),
            year=ExtractYear('tanggal_pemesanan')
        ).values('day', 'month', 'year').annotate(total_count=Count('SKU', distinct=True))
    else: 
        item_counts = Items.objects.count()
        approved_item_counts = Items.objects.filter(is_approved=True).count()
        grouped_counts = Items.objects.values('customer__nama_pt', 'pic__nama').annotate(count=Count('SKU')).order_by('-count', 'customer__nama_pt', 'pic__nama')
        category_counts = Items.objects.values('category__name').annotate(count=Count('SKU')).order_by('-count', 'category__name')
        sku_counts = Items.objects.annotate(
            day=ExtractDay('tanggal_pemesanan'),
            month=ExtractMonth('tanggal_pemesanan'),
            year=ExtractYear('tanggal_pemesanan')
        ).values('day', 'month', 'year').annotate(total_count=Count('SKU', distinct=True))

    category_counts_serialized = json.dumps(list(category_counts))
    sku_counts_serialized = json.dumps(list(sku_counts))

    print(sku_counts_serialized)
    
    context = {
        'form' : form,
        'item_counts' : item_counts,
        'approved_item_counts' : approved_item_counts,
        'grouped_counts' : grouped_counts,
        'category_counts' : category_counts_serialized,
        'sku_counts' : sku_counts_serialized,
    }

    return render(request, 'dashboardLS.html', context)

@login_required
def success(request):
    return render(request, 'success.html')


# -------------------- Common Functions --------------------#
# Adding entity (Customer and Supplier)
@login_required
def add_entity_view(request, entity_form, template_name, redirect_template, initial = None):
    entity_form_instance = entity_form(request.POST or None)
    if request.method == 'POST':
        form = entity_form(request.POST)
        if form.is_valid():
            form.save()
            return redirect(redirect_template)
    else:
        print(initial)
        if (initial):
            form = (entity_form(initial=initial))
            entity_form_instance = form
        else: 
            form = entity_form()

    return render(request, template_name, {'entity_form': entity_form_instance})


# Common add_entity function for adding alamat and pic
@login_required
def add_entity(request, entity_id, entity_model, form_class, template_name, entity_field, entity_form_field, initial_data=None, redirect_url = None):
    entity = get_object_or_404(entity_model, **{entity_field: entity_id})

    if request.method == 'POST':
        form = form_class(request.POST)
        if form.is_valid():
            setattr(form.instance, entity_form_field, entity)
            form.save()
            if redirect_url:
                return redirect(redirect_url)
    else:
        form = form_class(initial=initial_data)

    return render(request, template_name, {'entity': entity, 'form': form, 'entity_id': entity_id})


# Display entities for displaying tables
@login_required
def display_entities(request, entity_model, template_name):
    entities = entity_model.objects.all()
    return render(request, template_name, {'entities': entities})


# -------------------- Common Functions for Detail, Edit, and Delete -------------------- #
@login_required
def entity_detail(request, entity_model, entity_form, entity_id_field, entity_id, template_name, extra_context=None):
    entity = get_object_or_404(entity_model, **{entity_id_field: entity_id})
    form = entity_form(instance=entity)
    context = {'entity': entity, 'form': form, 'entity_id': entity_id}

    if extra_context:
        context.update(extra_context)

    return render(request, template_name, context)

@login_required
def edit_entity(request, entity_model, entity_form, entity_id_field, entity_id):
    entity = get_object_or_404(entity_model, **{entity_id_field: entity_id})

    if request.method == 'POST':
        form = entity_form(request.POST, instance=entity)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = entity_form(instance=entity)

    return render(request, 'edit_entity.html', {'form': form})

@login_required
def delete_entity(request, entity_model, entity_id_field, entity_id):
    entity = get_object_or_404(entity_model, **{entity_id_field: entity_id})

    if request.method == 'POST':
        entity.delete()
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})


# -------------------- Add Customer and Supplier Views -------------------- #
@login_required
@GA_required
def add_customer(request):
    return add_entity_view(request, CustomerForm, 'customer/add_cust.html', 'display_customer')

@login_required
@GA_required
def add_supplier(request):
    return add_entity_view(request, SupplierForm, 'supplier/add_supp.html', 'display_supplier')


# -------------------- Add Alamat and PIC -------------------- #
@login_required
@GA_required
def add_customer_pic(request, cust_id):
    redirect_url  = reverse('customer_detail', args=(cust_id,))
    return add_entity(request, cust_id, Customer, CustPICForms, 'pic/add_cust_pic.html', 'cust_id', 'customer_id', {'customer_id': cust_id}, redirect_url=redirect_url)

@login_required
@GA_required
def add_supplier_pic(request, supp_id):
    redirect_url  = reverse('supplier_detail', args=(supp_id,))
    return add_entity(request, supp_id, Supplier, SuppPICForms, 'pic/add_supp_pic.html', 'supp_id', 'supplier_id', {'supplier_id': supp_id}, redirect_url=redirect_url)


@login_required
def add_prospect_pic(request, prospect_id):
    redirect_url  = reverse('prospect_detail', args=(prospect_id,))
    return add_entity(request, prospect_id, Prospect, ProspectPICForms, 'pic/add_prospect_pic.html', 'prospect_id', 'prospect_id', {'prospect_id': prospect_id}, redirect_url=redirect_url)

@login_required
@GA_required
def add_customer_alamat(request, cust_id):
    redirect_url  = reverse('customer_detail', args=(cust_id,))
    return add_entity(request, cust_id, Customer, CustAlamatForms, 'alamat/add_customer_alamat.html', 'cust_id', 'customer_id', {'customer_id': cust_id}, redirect_url=redirect_url)

@login_required
@GA_required
def add_supplier_alamat(request, supp_id):
    redirect_url  = reverse('supplier_detail', args=(supp_id,))
    return add_entity(request, supp_id, Supplier, SuppAlamattForms, 'alamat/add_supplier_alamat.html', 'supp_id', 'supplier_id', {'supplier_id': supp_id}, redirect_url=redirect_url)

@login_required
def add_prospect_alamat(request, prospect_id):
    redirect_url  = reverse('prospect_detail', args=(prospect_id,))
    return add_entity(request, prospect_id, Prospect, ProspectAlamatForm, 'alamat/add_prospect_alamat.html', 'prospect_id', 'prospect_id', {'prospect_id': prospect_id}, redirect_url=redirect_url)


# -------------------- Edit and Delete Alamat and PIC -------------------- #
@login_required
@GA_required
def edit_customer_pic(request, pic_id):
    pic = get_object_or_404(CustomerPIC, id=pic_id)
    form = CustPICForms(request.POST or None, instance=pic)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('customer_detail', cust_id=pic.customer_id.pk)
    return render(request, 'pic/edit_customer_pic.html', {'form': form, 'pic': pic})

@login_required
@GA_required
def delete_customer_pic(request, pic_id):
    return delete_entity(request, CustomerPIC, 'id', pic_id)

@login_required
@GA_required
def edit_customer_alamat(request, alamat_id):
    alamat = get_object_or_404(CustomerAlamat, id=alamat_id)
    form = CustAlamatForms(request.POST or None, instance=alamat)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('customer_detail', cust_id=alamat.customer_id.pk)
    return render(request, 'alamat/edit_customer_alamat.html', {'form': form, 'alamat': alamat})

@login_required
@GA_required
def delete_customer_alamat(request, alamat_id):
    return delete_entity(request, CustomerAlamat, 'id', alamat_id)

# Supplier
@login_required
@GA_required
def edit_supplier_pic(request, pic_id):
    pic = get_object_or_404(SupplierPIC, id=pic_id)
    form = SuppPICForms(request.POST or None, instance=pic)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('supplier_detail', supp_id=pic.supplier_id.pk)
    return render(request, 'pic/edit_supplier_pic.html', {'form': form, 'pic': pic})

@login_required
@GA_required
def delete_supplier_pic(request, pic_id):
    return delete_entity(request, SupplierPIC, 'id', pic_id)

@login_required
@GA_required
def edit_supplier_alamat(request, alamat_id):
    alamat = get_object_or_404(SupplierAlamat, id=alamat_id)
    form = SuppAlamattForms(request.POST or None, instance=alamat)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('customer_detail', cust_id=alamat.supplier_id.pk)
    return render(request, 'alamat/edit_supplier_alamat.html', {'form': form, 'alamat': alamat})

@login_required
@GA_required
def delete_supplier_alamat(request, alamat_id):
    return delete_entity(request, SupplierAlamat, 'id', alamat_id)

# Prospect
@login_required
def edit_prospect_pic(request, pic_id):
    pic = get_object_or_404(ProspectPIC, id=pic_id)
    form = ProspectPICForms(request.POST or None, instance=pic)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('prospect_detail', prospect_id=pic.prospect_id.pk)
    return render(request, 'pic/edit_prospect_pic.html', {'form': form, 'pic': pic})

@login_required
def delete_prospect_pic(request, pic_id):
    return delete_entity(request, ProspectPIC, 'id', pic_id)


@login_required
def edit_prospect_alamat(request, alamat_id):
    alamat = get_object_or_404(ProspectAddress, id=alamat_id)
    form = ProspectAlamatForm(request.POST or None, instance=alamat)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('prospect_detail', prospect_id=alamat.prospect_id.pk)
    return render(request, 'alamat/edit_prospect_alamat.html', {'form': form, 'alamat': alamat})

@login_required
def delete_prospect_alamat(request, alamat_id):
    return delete_entity(request, ProspectAddress, 'id', alamat_id)


# -------------------- Add Item -------------------- #
@login_required
@GA_required
def add_item(request):
    # form_instance = ItemForm(request.POST or None)
    if request.method == 'POST':
        form = ItemForm(request.POST, request.FILES)
        if form.is_valid():
            item = form.save(commit=False)
            image = item.gambar
            if image:
                # Open the image
                img = Image.open(image)
                
                # Resize the image
                img = img.resize((100, 100))  # Change the dimensions as needed
                regex_pattern = r'/'
                replacement_string = '.'
                nama_cleaned = re.sub(regex_pattern, replacement_string, item.nama)
                curr_datetime = datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')
                # Save the resized image
                # image_name = f"{item.nama}.{image.name.split('.')[-1]}"
                # image_path = os.path.join(settings.MEDIA_ROOT, image_name)
                resized_image_name = f"media_{nama_cleaned}_{curr_datetime}.{image.name.split('.')[-1]}"  # Rename the file to avoid overwriting the original
                resized_image_path = os.path.join(settings.MEDIA_ROOT, resized_image_name)
                img.save(resized_image_path)
                

                # os.remove(image_path)

                item.gambar = resized_image_name
                print(item.gambar)
                item.save()
            return redirect('display_item')

    else:
        form = ItemForm()

    return render(request, 'item/add_item.html', {'item_form': form})


# -------------------- Display Tables -------------------- #
@login_required
@GA_required
def display_customer(request):
    return display_entities(request, Customer, 'customer/display_customer.html')

@login_required
@GA_required
def display_supplier(request):
    return display_entities(request, Supplier, 'supplier/display_supplier.html')

@login_required
@GA_required
def item_list(request):
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_column = request.GET.get('search_col')
    search_value = request.GET.get('search_val')

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    items = Items.objects.all()

    # Date range filtering
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
        items = items.filter(tanggal_pemesanan__range=[start_date, end_date])

    # Define the mapping outside the conditional block
    mapping = {
        '0': 'upload_type',
        '1': 'Tanggal',
        '2': 'tanggal_pemesanan',
        '3': 'customer__nama_pt',
        '4': 'pic__nama',
        '5': 'SKU',
        '6': 'nama',
        '7': 'category__name',
        '8': 'catatan',
        '9': 'quantity',
        '10': 'price',
        '13': 'is_approved',
    }

    # Column-based search
    if search_value:
        if search_column:
            field = mapping.get(search_column)
            if field:
                if search_column == '13':  # Handle the 'is_approved' column search
                    # Handle "Yes" and "No" filtering based on the approved status
                    if search_value.lower() == 'yes':
                        items = items.filter(is_approved=True)
                    elif search_value.lower() == 'no':
                        items = items.filter(is_approved=False)
                elif search_column in ['1', '2']:  # Handle date columns
                    try:
                        date_value = datetime.strptime(search_value, '%Y-%m-%d').date()
                        items = items.filter(**{f"{field}": date_value})
                    except ValueError:
                        items = items.none()
                else:
                    filter_kwargs = {f"{field}__icontains": search_value}
                    items = items.filter(**filter_kwargs)
            else:
                items = items.filter(Q(nama__icontains=search_value) | Q(SKU__icontains=search_value))
        else:
            items = items.filter(Q(nama__icontains=search_value) | Q(SKU__icontains=search_value))

    # Ordering
    order_column_index = int(request.GET.get('order[0][column]', 2))  # Default column to sort by is column 2 (tanggal_pemesanan)
    order_direction = request.GET.get('order[0][dir]', 'desc')  # Default direction is descending
    order_column = mapping.get(str(order_column_index), 'tanggal_pemesanan')
    if order_direction == 'asc':
        items = items.order_by(order_column)
    else:
        items = items.order_by(f'-{order_column}')

    # Pagination
    paginator = Paginator(items, length)
    page_number = (start // length) + 1
    page = paginator.get_page(page_number)

    # Serialize data
    data = []
    for item in page:
        item_detail_url = reverse('item_detail', args=[item.SKU])
        data.append([
            item.upload_type,
            item.Tanggal.strftime('%Y/%m/%d') if item.Tanggal else 'None',
            item.tanggal_pemesanan.strftime('%Y/%m/%d') if item.tanggal_pemesanan else 'None',
            str(item.customer) if item.customer else 'None',
            str(item.pic) if item.pic else 'None',
            item.SKU,
            item.nama,
            item.catatan if item.catatan else 'None',
            item.category.name if item.category else 'None',
            f"{item.quantity} {item.unit}",
            str(item.price),
            f'<img src="{item.gambar.url}" alt="{item.nama}" width="50" height="50" />' if item.gambar else 'None',
            ' '.join([f'<a href="{sumber.url}">{sumber.url[:20]}{"..." if len(sumber.url) > 20 else ""}</a>' for sumber in item.itemsumber_set.all() if sumber.url]) or 'None',
            'Yes' if item.is_approved else 'No',
            f'<a href="{item_detail_url}">View</a>'
        ])

    # Prepare response
    response = {
        'draw': draw,
        'recordsTotal': Items.objects.count(),
        'recordsFiltered': items.count(),
        'data': data,
    }

    return JsonResponse(response)

def display_item(request):
    return render(request, 'item/display_item.html')


def format_catatan(text):
    
    return text.replace("\n", "<br><br>")

def export_pdf_view(request):
    if request.method == "POST":
        try:
            # Parsing data JSON dari request body
            data = json.loads(request.body)
            selected_data = data.get("data", [])
            visible_columns = data.get("columns", [])

            ## DEBUGGING PURPOSE CHECKING CATATAN 
            # **Cari indeks kolom "Catatan"**
            # try:
            #     catatan_index = visible_columns.index("Catatan")  # Cari posisi kolom "Catatan"
            # except ValueError:
            #     catatan_index = -1  # Jika tidak ditemukan, set -1

            # # **Cek apakah kolom "Catatan" ada**
            # if catatan_index != -1:
            #     catatan_values = [row[catatan_index] for row in selected_data]  # Ambil hanya data di kolom "Catatan"
            #     print("Kolom Catatan:", catatan_values)  # Debug: tampilkan data
            # else:
            #     print("Kolom 'Catatan' tidak ditemukan dalam data yang dikirim.")




            if not selected_data or not visible_columns:
                return JsonResponse({"error": "Data tidak valid"}, status=400)

            # Cari indeks kolom "Catatan"
            try:
                catatan_index = visible_columns.index("Catatan")
            except ValueError:
                catatan_index = -1

            # Format hanya kolom "Catatan"
            if catatan_index != -1:
                for row in selected_data:
                    row[catatan_index] = format_catatan(row[catatan_index])

            # Simpan data ke session agar bisa diakses di export_pdf.html
            request.session["export_data"] = selected_data
            request.session["export_columns"] = visible_columns

            return JsonResponse({"pdf_url": reverse('export_pdf_page')})

        except json.JSONDecodeError:
            return JsonResponse({"error": "Format JSON tidak valid"}, status=400)

    # Untuk GET, render halaman PDF
    return render(request, "export_pdf.html")



# -------------------- Customer Functions -------------------- #
@login_required
@GA_required
def customer_detail(request, cust_id):
    customer_pics = CustomerPIC.objects.filter(customer_id=cust_id)
    customer_alamat = CustomerAlamat.objects.filter(customer_id=cust_id)
    extra_context = {'customer_pics':customer_pics, 'customer_alamat':customer_alamat}
    return entity_detail(request, Customer, CustomerForm, 'cust_id', cust_id, 'customer/customer_detail.html', extra_context)

@login_required
@GA_required
def edit_customer(request, cust_id):
    return edit_entity(request, Customer, CustomerForm, 'cust_id', cust_id)

@login_required
@GA_required
def delete_customer(request, cust_id):
    return delete_entity(request, Customer, 'cust_id', cust_id)


# -------------------- Customer Functions -------------------- #
@login_required
@GA_required
def supplier_detail(request, supp_id):
    supplier_pics = SupplierPIC.objects.filter(supplier_id=supp_id)
    supplier_alamat = SupplierAlamat.objects.filter(supplier_id=supp_id)
    extra_context = {'supplier_pics':supplier_pics, 'supplier_alamat':supplier_alamat}
    return entity_detail(request, Supplier, SupplierForm, 'supp_id', supp_id, 'supplier/supplier_detail.html', extra_context)

@login_required
@GA_required
def edit_supplier(request, supp_id):
    return edit_entity(request, Supplier, SupplierForm, 'supp_id', supp_id)

@login_required
@GA_required
def delete_supplier(request, supp_id):
    return delete_entity(request, Supplier, 'supp_id', supp_id)


# -------------------- Item Functions -------------------- #
@login_required
@GA_required
def item_detail(request, SKU):
    entity = get_object_or_404(Items, SKU=SKU)
    item_sumber = ItemSumber.objects.filter(item=SKU)
    change_logs = ItemChangeLog.objects.filter(item=entity)
    field_label_mapping = {
        'customer': 'Customer',
        'Tanggal': 'Tanggal Input',
        'tanggal_pemesanan': 'Tanggal Pemesanan',
        'nama': 'Nama Barang',
        'catatan': 'Catatan',
        'quantity': 'Kuantitas',
        'unit': 'Satuan',
        'price': 'Harga',
        'gambar': 'Gambar',
        'is_approved': 'Approved',
        'price currency': 'Mata Uang',
    }
    
    for log in change_logs:
        log.field_changed = field_label_mapping.get(log.field_changed, log.field_changed)

    context = {'entity':entity, 'item_sumber':item_sumber, 'form':ItemForm(instance=entity), 'change_logs':change_logs, 'field_label_mapping': field_label_mapping}
    return render(request, 'item/item_detail.html', context)

@login_required
@GA_required
def edit_item(request, SKU):
    entity = get_object_or_404(Items,SKU=SKU)
    entity_approved = entity.is_approved

    if request.method == 'POST':
        form = ItemForm(request.POST, request.FILES, instance=entity)
        if form.is_valid():
            # Check if a new image file is provided
            new_image = request.FILES.get('gambar')
            if new_image:
                # Process the new image file (similar to the logic in your add_item view)
                img = Image.open(new_image)
                img = img.resize((100, 100))
                resized_image_name = f"{entity.SKU}_resized.{new_image.name.split('.')[-1]}"
                resized_image_path = os.path.join(settings.MEDIA_ROOT, resized_image_name)
                img.save(resized_image_path)
                
                # Update the item's image field with the new image path
                form.instance.gambar = resized_image_name

            if request.user.groups.filter(name='Admin').exists():
                new_approved_raw = request.POST.get('is_approved', 'off')
                new_approved = new_approved_raw == "on"
                print(new_approved)
                if new_approved != entity.is_approved:
                    entity.is_approved = new_approved
            else:
                entity.is_approved = entity_approved

            # entity.is_approved = entity_approved

            form.save()

            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = ItemForm(instance=entity)

    return render(request, 'edit_item.html', {'form': form})

def get_pic_options(request):
    customer_id = request.GET.get('cust_id')
    pics = CustomerPIC.objects.filter(customer_id=customer_id)
    data = [{'id': pic.id, 'name': pic.name} for pic in pics]
    return JsonResponse(data, safe=False)

@login_required
@GA_required
def delete_item(request, SKU):
    entity = get_object_or_404(Items, SKU=SKU)
    image = entity.gambar
    image_str = str(image)
    image_path = os.path.join(settings.MEDIA_ROOT, image_str)
    print(image_path)
    os.remove(image_path)
    entity.delete()
    messages.success(request, 'Item deleted successfully')
    return redirect('/display_item')

@login_required
@GA_required
def approve_item(request, SKU):
    item = get_object_or_404(Items, SKU=SKU)
    item.is_approved = True
    item.save()
    # Redirect to the item list or any other appropriate view
    return redirect('display_item')

def get_customer_pics(request):
    if request.method == 'GET':
        customer_id = request.GET.get('customer_id')
        pics = CustomerPIC.objects.filter(customer_id=customer_id).values('id', 'nama')
        return JsonResponse(list(pics), safe=False)
    return JsonResponse ({'error': 'Invalid Request'})

def get_customer_by_pic(request):
    if request.method == 'GET':
        pic_id = request.GET.get('pic_id')
        # try: 
        customer_pic = CustomerPIC.objects.get(id=pic_id)
        customer_id = customer_pic.customer_id.cust_id
        # print(customer_pic)
        return JsonResponse({'customer_id' : customer_id})
        # except:
        #     return JsonResponse({'customer_id' : None})
    else: 
        return JsonResponse({'customer_id': None})


# -------------------- Item Sumber Functions -------------------- #
@login_required
@GA_required
def add_sumber(request, SKU):
    redirect_url  = reverse('item_detail', args=(SKU,))
    return add_entity(request, SKU, Items, SumberForm, 'item/add_sumber.html', 'SKU', 'item', redirect_url=redirect_url)

@login_required
@GA_required
def edit_sumber(request, sumber_id):
    sumber = get_object_or_404(ItemSumber, id=sumber_id)
    form = SumberForm(request.POST or None, instance=sumber)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('item_detail', SKU=sumber.item.pk)
    return render(request, 'item/edit_sumber.html', {'form': form, 'sumber': sumber})

# def edit_customer_pic(request, pic_id):
#     pic = get_object_or_404(CustomerPIC, id=pic_id)
#     form = CustPICForms(request.POST or None, instance=pic)

#     if request.method == 'POST':
#         if form.is_valid():
#             form.save()
#             return redirect('customer_detail', cust_id=pic.customer_id.pk)
#     return render(request, 'pic/edit_customer_pic.html', {'form': form, 'pic': pic})
@login_required
@GA_required
def delete_sumber(request, sumber_id):
    return delete_entity(request, ItemSumber, 'id', sumber_id)


# -------------------- Order Functions -------------------- #
# Add Purchase Order and Work Order
@login_required
@GA_required
def add_PO(request):
    return add_entity_view(request, PurchaseForm, 'order/add_PO.html', 'display_purchase')

@login_required
@GA_required
def add_WO(request):
    #  return add_entity_view(request, WorkForm, 'order/add_WO.html', 'display_work')
    if request.method == 'POST':
        form = WorkForm(request.POST)
        items_formset = WorkItemFormSet(request.POST)

        if form.is_valid() and items_formset.is_valid():
            work_order = form.save()
            items_formset = WorkItemFormSet(request.POST, instance=work_order)
            if items_formset.is_valid():
                # print("Valid Form")
                # for form in items_formset:
                #     form.instance.work_order = work_order
                items_formset.save()
                return redirect('display_work')
            else:
                print("Formset errors:", items_formset.errors)
                print("Formset data:", request.POST)  # Print formset data for debugging purposes
        else:
            print("Main form errors:", form.errors)
    else:
        form = WorkForm()
        items_formset = WorkItemFormSet()
    
    return render(request, 'order/add_WO.html', {'entity_form' : form, 'items_formset' : items_formset})

# Display Purchase Order and Work Order
@login_required
@Messenger_Forbidden
def display_purchase(request):
    return display_entities(request, PurchaseOrder, 'order/display_purchase.html')

@login_required
@Messenger_Forbidden
def display_work(request):
    entities = WorkOrder.objects.all().prefetch_related(
        Prefetch('workorderitems_set', queryset=WorkOrderItems.objects.select_related('item'))
    )
    items = WorkOrderItems.objects.all()
    return render(request,'order/display_work.html', {'entities': entities, 'items':items})

#  Detail of Purchase Order and Work Order
@login_required
@Messenger_Forbidden
def purchase_detail(request, id):
    entity = get_object_or_404(PurchaseOrder, id = id)

    if request.user.groups.filter(name='GA').exists() or request.user.groups.filter(name='Admin').exists():
            form = PurchaseForm(instance=entity)
    elif request.user.groups.filter(name='Accounting').exists():
        form = PurchaseFormNGA(instance=entity)
    context = {'entity': entity, 'form': form, 'entity_id': id}

    return render(request, 'order/purchase_detail.html', context)

@login_required
@Messenger_Forbidden
def work_detail(request, id):
    entity = get_object_or_404(WorkOrder, id = id)

    if request.user.groups.filter(name='GA').exists() or request.user.groups.filter(name='Admin').exists():
        form_class = WorkForm
    elif request.user.groups.filter(name='Accounting').exists():
        form_class = WorkFormNGA

    if request.method == 'POST':
        form = form_class(request.POST, instance=entity)
        items_formset = WorkItemFormSet(request.POST, instance=entity)
        
        if form.is_valid() and items_formset.is_valid():
            # print("Valid")
            form.save()
            for form in items_formset:
                print("form: ", form)
                form.save()
            items_formset.save()
            return redirect('display_work')
        else: 
            print("Not Valid")
    else:
        form = form_class(instance=entity)
        items_formset = WorkItemFormSet(instance=entity)

    context = {'entity': entity, 'form': form, 'items_formset':items_formset, 'entity_id': id}

    return render(request, 'order/work_detail.html', context)

# Edit Purchase Order and Work Order
@login_required
@Messenger_Forbidden
def edit_purchase(request, id):
    entity = get_object_or_404(PurchaseOrder, id = id)

    if request.method == 'POST':
        if request.user.groups.filter(name='GA').exists() or request.user.groups.filter(name='Admin').exists():
            form = PurchaseForm(request.POST, instance=entity)
        elif request.user.groups.filter(name='Accounting').exists():
            form = PurchaseFormNGA(request.POST, instance=entity)
        
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

    return render(request, 'order/purchase_detail.html', {'form': form})

@login_required
@Messenger_Forbidden
def edit_work(request, id):
    entity = get_object_or_404(WorkOrder, id = id)

    if request.method == 'POST':
        if request.user.groups.filter(name='GA').exists() or request.user.groups.filter(name='Admin').exists():
            form = WorkForm(request.POST, instance=entity)
        elif request.user.groups.filter(name='Accounting').exists():
            form = WorkFormNGA(request.POST, instance=entity)

        item_formset = WorkItemFormSet(request.POST, instance=entity)
        
        if form.is_valid() and item_formset.is_valid():
            form.save()
            for form in item_formset:
                if form.cleaned_data.get('delete'):
                    form.delete()
                else:
                    form.instance.entity = entity 
                    form.save()
                # print(form)
            item_formset.save()
            return JsonResponse({'success': True})
            # return redirect('display_work')
        else:
            print("Form Error:",  item_formset.errors)
            return JsonResponse({'success': False, 'errors': form.errors})

    return render(request, 'order/work_detail.html', {'form': form, 'item_formset': item_formset})

# Delete Purchase Order and Work Order
@login_required
@Messenger_Forbidden
def delete_purchase(request, id):
    return delete_entity(request, PurchaseOrder, 'id', id)

@login_required
@Messenger_Forbidden
def delete_work(request, id):
    return delete_entity(request, WorkOrder, 'id', id)

def get_customer_pics(request):
    if request.method == 'GET':
        customer_id = request.GET.get('customer_id')
        pics = CustomerPIC.objects.filter(customer_id=customer_id).values('id', 'nama')
        return JsonResponse(list(pics), safe=False)
    else:
        return JsonResponse ({'error': 'Invalid Request'})

def get_customer_by_pic(request):
    if request.method == 'GET':
        pic_id = request.GET.get('pic_id')
        # try: 
        customer_pic = CustomerPIC.objects.get(id=pic_id)
        customer_id = customer_pic.customer_id.cust_id
        # print(customer_pic)
        return JsonResponse({'customer_id' : customer_id})
        # except:
        #     return JsonResponse({'customer_id' : None})
    else: 
        return JsonResponse({'customer_id': None})

def get_customer_item(request):
    if request.method == 'GET':
        customer_id = request.GET.get('customer_id')
        items = Items.objects.filter(customer=customer_id, is_approved=True).values('SKU', 'nama', 'price')
        return JsonResponse(list(items), safe=False)
    else:
        return JsonResponse ({'error': 'Invalid Request'})

def get_item_details(request):
    if request.method == 'GET':
        item_id = request.GET.get('item_id')
        try:
            item = Items.objects.get(SKU=item_id)
            # Assuming 'price' is a field in your Item model
            item_details = {
                'SKU': item.SKU,
                'nama': item.nama,
                'price': str(item.price),
                'price_currency': item.price_currency,
                'quantity': item.quantity,
                'unit': item.unit,
            }
            return JsonResponse(item_details)
        except Items.DoesNotExist:
            return JsonResponse({'error': 'Item not found'}, status=404)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)


# -------------------- Login, Register, Logout Functions -------------------- #
# Login, Register, and Logout
def login_view(request):
    if request.method == 'POST':
        form = Login(request, request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            user = authenticate(request, username=email, password=password)
            if user is not None:
                login(request, user)
                return redirect('/')
    else:
        form = Login()
    
    form.fields['email'].widget.attrs.update({'class':'form-control'})
    # form.fields['username'].widget.attrs.update({'class': 'form-control'})
    form.fields['password'].widget.attrs.update({'class': 'form-control'})
    return render(request, 'accounts/login.html', {'form': form})

def register_view(request):
    if request.method == 'POST':
        form = Register(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/login')
    else:
        form = Register()
    return render(request, 'accounts/register.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('/login')


# -------------------- User Action Logs -------------------- #
@login_required
@Admin_Only
def user_action_logs(request):
    logs = UserActionLog.objects.all().order_by('-timestamp')[:100]  # Get the last 10 logs
    return render(request, 'logs.html', {'logs': logs})


# -------------------- Delivery Order -------------------- #
# Display calendar, and event addition/deletion functionality
@login_required
@Messenger_Only
def calendar(request):  
    all_events = Events.objects.all()
    messengers = Messenger.objects.all()
    request.session['num_forms'] = 1
    request.session.modified = True
    context = {
        "events":all_events,
        "messengers": messengers,
    }
    return render(request,'delivery/calendar.html',context)

@login_required
@Messenger_Only
def all_events(request):                                                                                                 
    all_events = Events.objects.all()                                                                                    
    out = []                                                                                                             

    for event in all_events: 
        event.start = event.start.astimezone(timezone.get_current_timezone())                                                                                            
        event.end = event.end.astimezone(timezone.get_current_timezone())                                                                                            
        out.append({     
            'title': event.title,                                                                                    
            'id': event.id,                                                                                              
            'start': event.start.strftime("%m/%d/%Y, %H:%M:%S"),                                                         
            'end': event.end.strftime("%m/%d/%Y, %H:%M:%S"),
            'messenger_id':event.messenger.id,      
            'messenger_color':event.messenger.color                                                       
        })                                                                                                               
                                                                                                                      
    return JsonResponse(out, safe=False) 
 
@login_required
@FO_Only
def add_event(request):
    start = request.GET.get("start", None)
    end = request.GET.get("end", None)
    title = request.GET.get("title", None)
    event = Events(name=str(title), start=start, end=end)
    event.save()
    data = {}
    return JsonResponse(data)
 
@login_required
@FO_Only
def update(request):
    start = request.GET.get("start", None)
    end = request.GET.get("end", None)
    title = request.GET.get("title", None)
    id = request.GET.get("id", None)
    event = Events.objects.get(id=id)
    event.start = start
    event.end = end
    event.title = title
    event.save()
    data = {}
    return JsonResponse(data)
 
@login_required
@FO_Only
def remove(request):
    id = request.GET.get("id", None)
    event = Events.objects.get(id=id)
    event.delete()
    data = {}
    return JsonResponse(data)

# Forms for adding delivery order, messenger, and vehicle
# Delivery form functionality
@login_required
@FO_Only
def delivery_form(request):
    max_forms = 3  # Maximum number of forms allowed
    
    if request.method == 'POST':
        num_forms = int(request.POST.get('num_forms', 1))  # Get the submitted number of forms

        forms = [DeliveryForm(request.POST, prefix=str(i)) for i in range(1, num_forms + 1)]
        
        if all(form.is_valid() for form in forms):
            # All forms are valid, process the data as needed
            for form in forms:
                # Save or process each form's data
                instance = form.save(commit=False)
                # Do additional processing if needed
                instance.save()

            request.session['num_forms'] = 1
            request.session.modified = True
            return redirect('calendar')  # Redirect to a success page
    else:
        start_param = request.GET.get('start')
        end_param = request.GET.get('end')
        # Set initial data for the forms based on start and end parameters
        initial_data = {'start': start_param, 'end': end_param}

        num_forms = int(request.session.get('num_forms', 1))
        # print(num_forms)
        forms = [DeliveryForm(initial=initial_data, prefix=str(i)) for i in range(1, num_forms + 1)]

    context = {'forms': forms, 'max_forms': max_forms}
    return render(request, 'delivery/delivery_form.html', context)

@login_required
@FO_Only
def update_num_forms(request):
    if request.method == 'POST':
        num_forms = int(request.POST.get('num_forms', 1))
        request.session['num_forms'] = num_forms
        request.session.modified = True
        
        return JsonResponse({'status': 'success', 'num_forms':num_forms})

    return JsonResponse({'status': 'error'})

# Display, edit, and delete delivery
@login_required
@FO_Only
def delivery_detail(request, id):
    return entity_detail(request, Events, DeliveryForm, "id", id, 'delivery/delivery_detail.html')

@login_required
@FO_Only
def edit_delivery(request, id):
    return edit_entity(request, Events, DeliveryForm, 'id', id)

@login_required
@FO_Only
def delete_delivery(request, id):
    return delete_entity(request, Events, 'id', id)

@login_required
@FO_Only
def display_delivery(request):
    return display_entities(request, Events, 'delivery/display_delivery.html')

# Adding messenger and vehicle
@login_required
def add_messenger(request):
    return add_entity_view(request, MessengerForm, 'delivery/add_messenger.html', 'calendar')

@login_required
def add_vehicle(request):
    return add_entity_view(request, VehicleForm, 'delivery/add_vehicle.html', 'calendar')

@login_required
def get_messenger(request):
    vehicle_id = request.GET.get('vehicle')
    print(vehicle_id)

    if vehicle_id:
        data = {}
        vehicle = Vehicle.objects.get(pk=vehicle_id)
        data['messenger'] = vehicle.messenger.id

        return JsonResponse(data)
    else:
        data = {}

    return JsonResponse(data)


# -------------------- Error Pages -------------------- #
def forbidden(request):
    return render(request, 'forbidden.html')

def page_not_found(request, exception):
    return render(request, '404.html', status=404)


# -------------------- Upload CSV/Excel -------------------- #
def upload_csv(request):
    if request.method == 'POST' and request.FILES['csv_file']:
        csv_file = request.FILES['csv_file']
        decoded_file = csv_file.read().decode('utf-8').splitlines()
        csv_reader = csv.DictReader(decoded_file)

        success_count = 0
        error_messages = []

        for row_number, row in enumerate(csv_reader, start=1):
            try:
                # Get or create the Category instance based on the provided category name
                category_name = row['category']
                category_instance, _ = Category.objects.get_or_create(name=category_name)

                # Check if image URL is provided in the CSV row
                if 'gambar' in row and row['gambar']:
                    image_url = row['gambar']
                    # Download the image from the URL
                    response = requests.get(image_url)
                    if response.status_code == 200:
                        # Create a file-like object from the image content
                        image_content = response.content
                        # Use the image URL's filename as the uploaded file name
                        image_filename = image_url.split('/')[-1]
                        # Create a SimpleUploadedFile object with the image content
                        uploaded_image = SimpleUploadedFile(image_filename, image_content)

                # Create the Items object with the retrieved Category instance
                item_data = {
                    'nama': row['nama'],
                    'category': category_instance,
                    'quantity': row['quantity'],
                    'unit': row['unit'],
                    'price': row['price'],
                    'price_currency' : row['price_currency'],
                    'upload_type': 'bulk'
                }
                if 'gambar' in row and row['gambar']:
                    item_data['gambar'] = uploaded_image

                item = Items.objects.create(**item_data)
                # item.upload_type = "bulk"
                success_count += 1
            except Exception as e:
                error_messages.append(f"Error in row {row_number}: {str(e)}")

        if error_messages:
            return JsonResponse({'success': False, 'errors': error_messages,'message':f"{success_count} items imported successfully."})
        else:
            return JsonResponse({'success': True, 'message': f"{success_count} items imported successfully."})

    return render(request, 'item/upload_csv.html')

@login_required
@GA_required
def upload_excel(request):
    error_message = []
    processed_items = []
    categories = Category.objects.all()
    customers = Customer.objects.all()
    customer_pics = CustomerPIC.objects.all()

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES['excel_file']
                wb = openpyxl.load_workbook(excel_file)
                worksheet = wb.active

                if worksheet.max_row < 2:
                    raise ValueError("File Excel tidak memiliki data yang cukup!")

                # Ambil header kolom
                header_row = next(worksheet.iter_rows(values_only=True))
                column_titles = [str(cell).strip().lower() for cell in header_row]

                # Mapping kolom
                column_indices = {
                    col: column_titles.index(col) if col in column_titles else None
                    for col in [
                        'tanggal_pemesanan', 'nama', 'catatan', 'category', 'customer',
                        'quantity', 'unit', 'price', 'price_currency', 'jenis_sumber',
                        'link', 'telp_sumber', 'email_sumber', 'nama_sumber', 'pic', 'gambar'
                    ]
                }

                # Cek jika kolom wajib tidak ada
                required_columns = ['nama', 'category', 'customer', 'quantity', 'unit', 'price']
                missing_columns = [col for col in required_columns if column_indices[col] is None]

                if missing_columns:
                    raise ValueError(f"Kolom wajib berikut tidak ditemukan di file Excel: {', '.join(missing_columns)}")

                # Temukan baris pertama yang memiliki "gambar"
                gambar_row_index = None
                for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
                    if 'gambar' in [str(cell).strip().lower() for cell in row]:
                        gambar_row_index = row_index
                        break

                if gambar_row_index is None:
                    raise ValueError("Kolom 'gambar' tidak ditemukan dalam file Excel!")

                # Load image dari worksheet
                image_loader = SheetImageLoader(worksheet)

                for row_index, row in enumerate(worksheet.iter_rows(min_row=gambar_row_index + 2, values_only=True)):
                    try:
                        # Ensure category is not empty
                        category_name = row[column_indices['category']] if column_indices['category'] is not None else None
                        if not category_name:
                            raise ValueError(f"Category is missing for row {row_index + 2}")

                        category_instance, _ = Category.objects.get_or_create(name=category_name)

                        # Handle customer and PIC
                        customer_instance, _ = Customer.objects.get_or_create(nama_pt=row[column_indices['customer']])
                        pic_instance, _ = CustomerPIC.objects.get_or_create(
                            customer_id=customer_instance, nama=row[column_indices['pic']]
                        )

                        # Set the order date to the current date if it's missing
                        tanggal_pemesanan = row[column_indices['tanggal_pemesanan']] if column_indices['tanggal_pemesanan'] is not None else datetime.now().strftime('%Y-%m-%d')

                        # Load image if there's one
                        filename = ""
                        if column_indices['gambar'] is not None:
                            image_cell = chr(65 + column_titles.index('gambar')) + str(row_index + 2)
                            try:
                                image = image_loader.get(image_cell)
                                image = image.resize((100, 100), Image.Resampling.LANCZOS)

                                item_name = re.sub(r'[\\/:"*?<>|\']', '-', row[column_indices['nama']])
                                upload_date = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                                filename = f"media_bulk_{item_name}_{upload_date}_{row_index}.png"
                                image_path = os.path.join(settings.MEDIA_ROOT, filename)

                                os.makedirs(settings.MEDIA_ROOT, exist_ok=True)  # Ensure the media folder exists
                                image.save(image_path)
                            except Exception as img_error:
                                print(f"Error loading image at {image_cell}: {img_error}")
                                filename = ""
                        # print(filename)
                        # Create the Items instance
                        instance = Items.objects.create(
                            tanggal_pemesanan=tanggal_pemesanan,
                            nama=row[column_indices['nama']],
                            catatan=row[column_indices['catatan']] if column_indices['catatan'] is not None else '',
                            category=category_instance,
                            customer=customer_instance,
                            pic=pic_instance,
                            quantity=row[column_indices['quantity']],
                            unit=row[column_indices['unit']],
                            price=row[column_indices['price']],
                            price_currency=row[column_indices['price_currency']] if column_indices['price_currency'] is not None else 'IDR',
                            gambar=filename,
                            upload_type="bulk"
                        )

                        # Create the associated ItemSumber
                        ItemSumber.objects.create(
                            item=instance,
                            jenis_sumber=row[column_indices['jenis_sumber']] if column_indices['jenis_sumber'] is not None else 'Online Store',
                            nama_perusahaan=row[column_indices['nama_sumber']] if column_indices['nama_sumber'] is not None else 'Unknown',
                            telp=row[column_indices['telp_sumber']] if column_indices['telp_sumber'] is not None else '',
                            email=row[column_indices['email_sumber']] if column_indices['email_sumber'] is not None else '',
                            url=row[column_indices['link']] if column_indices['link'] is not None else ''
                        )

                        processed_items.append(instance)
                    except Exception as row_error:
                        print(f"Error processing row {row_index + 2}: {row_error}")
                return render(request, 'item/upload_excel.html', {
                    'form': form,
                    'categories': categories,
                    'customers': customers,
                    'customerPIC': customer_pics,
                    'processed_items': processed_items
                })

            except openpyxl.utils.exceptions.InvalidFileException:
                error_message.append("Format file tidak valid. Harap unggah file Excel yang benar.")
            except ValueError as e:
                error_message.append(str(e))
        else:
            error_message.append("Data form tidak valid. Harap periksa kembali input Anda.")

    else:
        form = ExcelUploadForm()

    return render(request, 'item/upload_excel.html', {
        'form': form,
        'categories': categories,
        'customers': customers,
        'customerPIC': customer_pics,
        'error_message': error_message
    })

# -------------------- Delete multiple items -------------------- #
@login_required
def delete_selected_rows(request, model, key):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_ids[]')  # Assuming you're sending an array of selected IDs
        try:
            selected_items = model.objects.filter(**{f'{key}__in': selected_ids})

            # Additional logic for image deletion if applicable
            if hasattr(model, 'gambar'):
                for item in selected_items:
                    image_path = os.path.join(settings.MEDIA_ROOT, str(item.gambar))
                    if os.path.exists(image_path):
                        os.remove(image_path)
                    else:
                        print(f"Image file not found: {image_path}")

            selected_items.delete()  # Delete the selected rows from the database
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

@GA_required
def delete_selected_rows_item(request):
    return delete_selected_rows(request, Items, 'SKU')

def delete_selected_rows_cust(request):
    return delete_selected_rows(request, Customer, 'cust_id')

def delete_selected_rows_supp(request):
    return delete_selected_rows(request, Supplier, 'supp_id')

def delete_selected_rows_PO(request):
    return delete_selected_rows(request, PurchaseOrder, 'id')

def delete_selected_rows_WO(request):
    return delete_selected_rows(request, WorkOrder, 'id')

def delete_selected_rows_delivery(request):
    return delete_selected_rows(request, Events, 'id')

def delete_selected_rows_logbook(request):
    return delete_selected_rows(request, LogBook, 'id')

def delete_selected_rows_prospect(request):
    return delete_selected_rows(request, Prospect, 'prospect_id')


# -------------------- Approve Items -------------------- #
@login_required
@Admin_Only
def approve_selected_rows(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_ids[]')  # Assuming you're sending an array of selected IDs
        try:
            selected_items = Items.objects.filter(**{f'{"SKU"}__in': selected_ids})
            selected_items.update(is_approved = True)  # Delete the selected rows from the database

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})


# -------------------- Address Data -------------------- #
@login_required
@FO_Only
def add_additional_address(request):
    if request.method == 'POST':
        form = AdditionalAddressForm(request.POST)
        if form.is_valid():
            form.save()  # This saves the form data to the database
            # return redirect('success')
    else:
        form = AdditionalAddressForm()
    
    return render(request, 'delivery/add_delivery_address.html', {'form': form})

@login_required
@FO_Only
def get_location_data(request):
    # Fetch data from your database or any other source
    delivery_addresses = DeliveryAddresses.objects.all()
    data = serializers.serialize('json', delivery_addresses)
    return JsonResponse(data, safe=False)


# -------------------- Log Book Basics -------------------- #
@login_required
@FO_Only
def add_log(request):
    start_param = request.GET.get('start')
    end_param = request.GET.get('end')
    initial_data = {'start': start_param, 'end': end_param}
    return add_entity_view(request, LogBookForm, 'log_book/add_log.html', 'log_book', initial=initial_data)

@login_required
@FO_Only
def log_book(request):  
    all_events = LogBook.objects.all()
    context = {
        "events":all_events,
    }
    return render(request,'log_book/log-book.html',context)

@login_required
@FO_Only
def display_log(request):
    return display_entities(request, LogBook, 'log_book/display_log.html')
                            
@login_required
@FO_Only
def lb_all_events(request):                                                                                                 
    all_events = LogBook.objects.all()                                                                                    
    out = []                                                                                                             

    for event in all_events: 
        event.start = event.start.astimezone(timezone.get_current_timezone())                                                                                            
        event.end = event.end.astimezone(timezone.get_current_timezone())                                                                                            
        out.append({     
            'nama': event.nama,                                                                                    
            'id': event.id,                                                                                              
            'start': event.start.strftime("%m/%d/%Y, %H:%M:%S"),                                                         
            'end': event.end.strftime("%m/%d/%Y, %H:%M:%S"),                                                       
        })                                                                                                               
                                                                                                                      
    return JsonResponse(out, safe=False) 
 
@login_required
@FO_Only
def lb_add_event(request):
    start = request.GET.get("start", None)
    end = request.GET.get("end", None)
    nama = request.GET.get("nama", None)
    event = LogBook(name=str(nama), start=start, end=end)
    event.save()
    data = {}
    return JsonResponse(data)
 
@login_required
@FO_Only
def lb_update(request):
    start = request.GET.get("start", None)
    end = request.GET.get("end", None)
    nama = request.GET.get("nama", None)
    id = request.GET.get("id", None)
    event = LogBook.objects.get(id=id)
    event.start = start
    event.end = end
    event.nama = nama
    event.save()
    data = {}
    return JsonResponse(data)
 
@login_required
@FO_Only
def lb_remove(request):
    id = request.GET.get("id", None)
    event = LogBook.objects.get(id=id)
    event.delete()
    data = {}
    return JsonResponse(data)

@login_required
@FO_Only
def log_detail(request, id):
    return entity_detail(request, LogBook, LogBookForm, "id", id, 'log_book/log_detail.html')

@login_required
@FO_Only
def edit_log(request, id):
    return edit_entity(request, LogBook, LogBookForm, 'id', id)

@login_required
@FO_Only
def delete_log(request, id):
    return delete_entity(request, LogBook, 'id', id)


# -------------------- Dependable Alamat -------------------- #
def get_kota(request):
    if request.method == 'GET':
        province_id = request.GET.get('province_id')
        cities = Kota.objects.filter(provinsi_id=province_id).values('id', 'name')
        return JsonResponse(list(cities), safe=False)
    return JsonResponse({'error': 'Invalid request'})

def get_kecamatan(request):
    if request.method == 'GET':
        city_id = request.GET.get('city_id')
        district = Kecamatan.objects.filter(kota_id=city_id).values('id', 'name')
        return JsonResponse(list(district), safe=False)
    return JsonResponse({'error': 'Invalid request'})

def get_kelurahan(request):
    if request.method == 'GET':
        district_id = request.GET.get('district_id')
        village = Kelurahan.objects.filter(kecamatan_id=district_id).values('id', 'name')
        return JsonResponse(list(village), safe=False)
    return JsonResponse({'error': 'Invalid request'})

def get_region_details(request):
    region_id = request.GET.get('region_id')

    if region_id:
        data = {}
        if len(region_id) == 4:  # Kota ID
            kota = Kota.objects.get(pk=region_id)
            data['kota_id'] = kota.id
            data['provinsi_id'] = kota.provinsi_id.id
        elif len(region_id) == 7:  # Kecamatan ID
            kecamatan = Kecamatan.objects.get(pk=region_id)
            data['kecamatan_id'] = kecamatan.id
            data['kota_id'] = kecamatan.kota_id.id
            data['provinsi_id'] = kecamatan.kota_id.provinsi_id.id
        elif len(region_id) == 10:  # Kelurahan ID
            kelurahan = Kelurahan.objects.get(pk=region_id)
            data['kelurahan_id'] = kelurahan.id
            data['kecamatan_id'] = kelurahan.kecamatan_id.id
            data['kota_id'] = kelurahan.kecamatan_id.kota_id.id
            data['provinsi_id'] = kelurahan.kecamatan_id.kota_id.provinsi_id.id

        return JsonResponse(data)
    else:
        data = {}

    return JsonResponse(data)

def get_kode_pos(request):
    kelurahan_id = request.GET.get('kelurahan_id')
    if kelurahan_id:
        try:
            kelurahan = KodePos.objects.get(kelurahan_id=kelurahan_id)
            kode_pos = kelurahan.kode_pos
            return JsonResponse({'kode_pos': kode_pos})
        except ObjectDoesNotExist:
            return JsonResponse({'kode_pos': ''})
    return JsonResponse({'error': 'Invalid Kelurahan ID'}, status=400)


# -------------------- Prospect -------------------- #
@login_required
def display_prospect(request):
    return display_entities(request, Prospect, 'prospect/display_prospect.html')

@login_required
def add_prospect(request):
    if request.method == 'POST':
        form = ProspectForm(request.POST)
        if form.is_valid():
            # Create a Prospect instance but do not save it yet
            prospect = form.save(commit=False)
            # Set the creator field to the current authenticated user
            prospect.in_charge = request.user
            # Now save the Prospect instance with the creator assigned
            prospect.save()
            return redirect('display_prospect')  # Redirect to a success URL
    else:
        form = ProspectForm()
    return render(request, 'prospect/add_prospect.html', {'entity_form': form})

@login_required
def prospect_detail(request, prospect_id):
    prospect_pics = ProspectPIC.objects.filter(prospect_id=prospect_id)
    prospect_alamat = ProspectAddress.objects.filter(prospect_id=prospect_id)
    extra_context = {'prospect_pics':prospect_pics, 'prospect_alamat':prospect_alamat}
    return entity_detail(request, Prospect, ProspectForm, 'prospect_id', prospect_id, 'prospect/prospect_detail.html', extra_context)

@login_required
def edit_prospect(request, prospect_id):
    return edit_entity(request, Prospect, ProspectForm, 'prospect_id', prospect_id)

@login_required
def delete_prospect(request, prospect_id):
    return delete_entity(request, Prospect, 'prospect_id', prospect_id)

# TICKETING FOR PROSPECT
@login_required
def prospect_ticket(request, prospect_id):
    prospect = get_object_or_404(Prospect, prospect_id=prospect_id)
    prospect_tickets = ProspectTicket.objects.filter(prospect_id=prospect).order_by('-date')

    sorted_tickets = sorted(prospect_tickets, key=lambda x:(not x.open))

    for prospect_ticket in sorted_tickets:
        prospect_ticket.sorted_logs = prospect_ticket.ticketlog_set.order_by('-date')
    
    ticket_log_form = TicketLogForm()

    if request.method == "POST":
        entity = get_object_or_404(ProspectTicket, **{'id': request.POST.get('ticket_id')})
        ticket_log_form = TicketLogForm(request.POST)
        if ticket_log_form.is_valid():
            setattr(ticket_log_form.instance, "ticket_id", entity)
            # ticket_log_form.instance.ticket_id = request.POST.get('ticket_id')
            ticket_log_form.save()
            return redirect(request.path)

    context = {'prospect': prospect, 'prospect_id': prospect_id, 'prospect_tickets': sorted_tickets, 'ticket_log_form': ticket_log_form,}

    return render(request, 'prospect/prospect_ticket.html', context)


@login_required
def add_prospect_ticket(request, prospect_id):
    redirect_url  = reverse('prospect_ticket', args=(prospect_id,))
    return add_entity(request, prospect_id, Prospect, ProspectTicketForm, 'prospect/add_prospect_ticket.html', 'prospect_id', 'prospect_id', {'prospect_id': prospect_id}, redirect_url=redirect_url)

@login_required
def edit_prospect_ticket(request, log_id):
    log = get_object_or_404(ProspectTicket, id=log_id)
    form = ProspectTicketForm(request.POST or None, instance=log)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('prospect_ticket', prospect_id=log.prospect_id.pk)
    return render(request, 'prospect/edit_prospect_ticket.html', {'form': form, 'log': log})

@login_required
def delete_prospect_ticket(request, log_id):
    return delete_entity(request, ProspectTicket, 'id', log_id)

# LOG FOR TICKET
@login_required
def add_ticket_log(request, prospect_id):
    entity = get_object_or_404(ProspectTicket, **{'id': prospect_id})
    redirect_url = reverse('prospect_ticket', args=(entity.prospect_id.pk,))
    return add_entity(request, prospect_id, ProspectTicket, TicketLogForm, 'prospect/log/add_ticket_log.html', 'id', 'ticket_id', {'ticket_id': prospect_id}, redirect_url=redirect_url)

@login_required
def edit_ticket_log(request, log_id):
    log = get_object_or_404(TicketLog, id=log_id)
    form = TicketLogForm(request.POST or None, instance=log)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('prospect_ticket', prospect_id=log.ticket_id.prospect_id.pk)
    return render(request, 'prospect/log/edit_ticket_log.html', {'form': form, 'log': log})

@login_required
def delete_ticket_log(request, log_id):
    return delete_entity(request, TicketLog, 'id', log_id)

# CONVERT TO CUSTOMER
@login_required
def convert_to_customer(request, prospect_id):
    prospect = get_object_or_404(Prospect, pk=prospect_id)

    # Create Customer object
    customer = Customer(
        nama_pt=prospect.nama,
        telp=prospect.telp,
        # Add other fields as needed
    )
    customer.save()

    # Create CustomerPIC object if ProspectPIC exists
    prospect_pic = prospect.prospectpic_set.first()
    if prospect_pic:
        customer_pic = CustomerPIC(
            customer_id=customer,
            nama=prospect_pic.nama,
            email=prospect.email,
            telp=prospect.telp,
            Role=prospect_pic.Role
        )
        customer_pic.save()

    # Create CustomerAlamat object if ProspectAddress exists
    prospect_address = prospect.prospectaddress_set.first()
    if prospect_address:
        customer_address = CustomerAlamat(
            customer_id=customer,
            type='pengiriman',  # Assuming a default type
            kode_pos=prospect_address.kode_pos,
            provinsi=prospect_address.provinsi,
            kota=prospect_address.kota,
            kecamatan=prospect_address.kecamatan,
            kelurahan=prospect_address.kelurahan,
            detail=prospect_address.detail
        )
        customer_address.save()

    # Deactivate Prospect
    prospect.open = False
    prospect.is_customer = True
    prospect.save()

    return redirect('display_prospect')  # Redirect to customer detail page